"""openpyxl 기반 엑셀 파싱/쓰기 유틸"""

import os
import shutil
import tempfile
from copy import copy
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


def col_letter_to_idx(letter: str) -> int:
    """열 문자 → 1-based 인덱스. 'A'→1, 'B'→2"""
    return column_index_from_string(letter.upper())


def is_formula_cell(cell) -> bool:
    """수식 셀 여부"""
    return isinstance(cell.value, str) and cell.value.startswith("=")


def is_empty_cell(cell) -> bool:
    """빈 셀 여부"""
    return cell.value is None or (isinstance(cell.value, str) and cell.value.strip() == "")


def get_sheet_names(file_path: str) -> list[str]:
    """엑셀 파일의 시트 이름 목록"""
    wb = load_workbook(file_path, read_only=True, data_only=True)
    names = wb.sheetnames
    wb.close()
    return names


def find_sheet(file_path: str, candidates: list[str]) -> str | None:
    """시트 이름 후보 목록에서 첫 번째 매칭 시트 반환"""
    sheet_names = get_sheet_names(file_path)
    for candidate in candidates:
        for name in sheet_names:
            if candidate.lower() in name.lower():
                return name
    return None


def parse_detail_sheet(
    file_path: str,
    sheet_name: str,
    column_mapping: dict,
    data_start_row: int,
) -> list[dict]:
    """세부내역 시트 파싱 → 행 리스트 반환"""
    wb = load_workbook(file_path, data_only=True)
    ws = wb[sheet_name]

    col_map = {}
    for field, letter in column_mapping.items():
        if field == "data_start_row":
            continue
        col_map[field] = col_letter_to_idx(letter)

    rows = []
    for row_num in range(data_start_row, ws.max_row + 1):
        item_name_col = col_map.get("item_name")
        if not item_name_col:
            continue

        cell = ws.cell(row=row_num, column=item_name_col)
        if is_empty_cell(cell):
            continue

        item_name = str(cell.value).strip()
        # 소계/합계 행 건너뛰기
        if item_name in ("합계", "소계", "총합계", "총계", "TOTAL", "SUB TOTAL"):
            continue

        row_data = {
            "item_name_raw": item_name,
            "source_sheet": sheet_name,
            "source_row": row_num,
        }

        for field, col_idx in col_map.items():
            if field == "item_name":
                continue
            val = ws.cell(row=row_num, column=col_idx).value
            if val is not None and isinstance(val, str):
                val = val.strip() if val.strip() != "" else None
            row_data[field] = val

        rows.append(row_data)

    wb.close()
    return rows


def write_results_to_excel(
    source_path: str,
    output_path: str,
    sheet_name: str,
    column_mapping: dict,
    items: list[dict],
    data_start_row: int,
) -> str:
    """원본 복사 후 결과 되쓰기. 수식 셀 skip."""
    shutil.copy2(source_path, output_path)
    wb = load_workbook(output_path)
    ws = wb[sheet_name]

    col_map = {}
    for field, letter in column_mapping.items():
        if field == "data_start_row":
            continue
        col_map[field] = col_letter_to_idx(letter)

    # source_row → item 매핑
    item_by_row = {item["source_row"]: item for item in items}

    for row_num in range(data_start_row, ws.max_row + 1):
        if row_num not in item_by_row:
            continue
        item = item_by_row[row_num]
        for field, col_idx in col_map.items():
            if field in ("item_name", "data_start_row"):
                continue
            cell = ws.cell(row=row_num, column=col_idx)
            if is_formula_cell(cell):
                continue  # 수식 셀 보존
            if field in item and item[field] is not None:
                cell.value = item[field]

    wb.save(output_path)
    wb.close()
    return output_path


def write_classification_sheet(
    output_path: str,
    items: list[dict],
    sheet_name: str = "공정분류결과",
) -> str:
    """분류 결과를 별도 시트로 추가. 원본 시트는 보존."""
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = load_workbook(output_path)

    # 기존 시트 있으면 제거 후 재생성
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    headers = [
        "원본행",
        "품명(원본)",
        "표준품명",
        "대공종",
        "세부공종",
        "규격",
        "단위",
        "수량",
        "단가",
        "금액",
        "신뢰도",
        "확인여부",
    ]
    ws.append(headers)

    # 헤더 스타일
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    center = Alignment(horizontal="center", vertical="center")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # 정렬: source_row 기준
    sorted_items = sorted(items, key=lambda x: x.get("source_row") or 0)

    for item in sorted_items:
        confirmed = "✓" if item.get("confirmed_at") else ("검토필요" if item.get("review_flag") else "자동")
        ws.append([
            item.get("source_row"),
            item.get("item_name_raw"),
            item.get("item_name_std"),
            item.get("process_major"),
            item.get("process_minor"),
            item.get("spec"),
            item.get("unit"),
            item.get("qty"),
            item.get("unit_price"),
            item.get("amount"),
            round(item["confidence"], 2) if item.get("confidence") is not None else None,
            confirmed,
        ])

    # 열 폭 자동
    widths = [8, 32, 28, 14, 16, 20, 8, 10, 12, 14, 10, 10]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    # 필터 + 첫 행 고정
    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    wb.save(output_path)
    wb.close()
    return output_path


def save_upload_to_temp(file_bytes: bytes, filename: str) -> str:
    """업로드 파일을 임시 경로에 저장"""
    temp_dir = tempfile.mkdtemp(prefix="jh_estimate_")
    file_path = os.path.join(temp_dir, filename)
    with open(file_path, "wb") as f:
        f.write(file_bytes)
    return file_path


def make_output_path(source_path: str) -> str:
    """원본 경로 기반 출력 파일 경로 생성"""
    base, ext = os.path.splitext(source_path)
    return f"{base}_result{ext}"


def create_pdf_result_excel(items: list, original_filename: str) -> str:
    """PDF 업로드 결과용 신규 Excel 파일 생성 → 임시 파일 경로 반환"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "견적 결과"

    headers = ["항목명(원본)", "표준명", "규격", "단위", "수량", "단가", "금액", "대공정", "소공정", "신뢰도", "검토"]
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row, item in enumerate(items, 2):
        ws.cell(row=row, column=1, value=item.get("item_name_raw", ""))
        ws.cell(row=row, column=2, value=item.get("item_name_std", ""))
        ws.cell(row=row, column=3, value=item.get("spec", ""))
        ws.cell(row=row, column=4, value=item.get("unit", ""))
        ws.cell(row=row, column=5, value=item.get("qty"))
        ws.cell(row=row, column=6, value=item.get("unit_price"))
        ws.cell(row=row, column=7, value=item.get("amount"))
        ws.cell(row=row, column=8, value=item.get("process_major", ""))
        ws.cell(row=row, column=9, value=item.get("process_minor", ""))
        ws.cell(row=row, column=10, value=item.get("confidence"))
        ws.cell(row=row, column=11, value="요검토" if item.get("review_flag") else "")

    tmp = tempfile.mkdtemp(prefix="jh_pdf_result_")
    stem = os.path.splitext(os.path.basename(original_filename))[0]
    out_path = os.path.join(tmp, f"{stem}_result.xlsx")
    wb.save(out_path)
    return out_path
