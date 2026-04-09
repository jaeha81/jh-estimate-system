"""
데모용 샘플 엑셀 파일 생성 스크립트

한샘리하우스 34평 풀 인테리어 견적서 양식 생성
- 세부내역서 시트, 데이터 시작 7행
- 18개 품목 (자동분류 13개 + review_flag 5개)
- 총 견적액: 22,597,500원

사용법:
  pip install openpyxl
  python scripts/create_demo_excel.py
  # 출력: docs/demo/한샘_리하우스_34평_견적서_샘플.xlsx
"""

from pathlib import Path
import openpyxl
from openpyxl.styles import (
    Font, Alignment, PatternFill, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── 출력 경로 ──────────────────────────────────────────────────
ROOT = Path(__file__).parent.parent
OUTPUT_DIR = ROOT / "docs" / "demo"
OUTPUT_FILE = OUTPUT_DIR / "한샘_리하우스_34평_견적서_샘플.xlsx"

# ── 스타일 정의 ────────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
SECTION_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
WHITE_FILL = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
LIGHT_GRAY = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

HEADER_FONT = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(name="맑은 고딕", bold=True, size=14)
SECTION_FONT = Font(name="맑은 고딕", bold=True, size=10, color="1F3864")
BODY_FONT = Font(name="맑은 고딕", size=9)
TOTAL_FONT = Font(name="맑은 고딕", bold=True, size=10)

THIN = Side(border_style="thin", color="BFBFBF")
MEDIUM = Side(border_style="medium", color="4F81BD")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_BORDER = Border(left=MEDIUM, right=MEDIUM, top=MEDIUM, bottom=MEDIUM)

NUM_FORMAT = "#,##0"
AREA_FORMAT = "#,##0.00"

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")

# ── 18개 품목 데이터 ────────────────────────────────────────────
# (번호, 품명, 규격, 단위, 수량, 재료비단가, 노무비단가, 경비)
# 단가 = 재료비+노무비+경비, 금액 = 수량×단가 (수식으로 처리)

ITEMS = [
    # 자동 분류 13개
    (1,  "석고보드 9.5T 천장",   "9.5T",          "㎡",   85,   10000, 7000,  1000),
    (2,  "실크 도배",            "합지실크",        "㎡",  320,    4000, 4000,   500),
    (3,  "강마루 12T",           "12T 강화마루",    "㎡",  112,   28000,15000,  2000),
    (4,  "욕실 타일 300×300",    "300×300 자기질",  "㎡",   28,   30000,22000,  3000),
    (5,  "주방 타일 200×300",    "200×300",         "㎡",   12,   22000,18000,  2000),
    (6,  "도어 교체",            "ABS도어 900×2100","EA",    8,  150000,100000,30000),
    (7,  "수성페인트 2회",        "수성",            "㎡",  180,    5000, 6000,  1000),
    (8,  "기존 바닥재 철거",      "-",               "㎡",  112,    3000, 4500,   500),
    (9,  "폐기물 처리",           "5톤",             "회",    2,  150000,170000,30000),
    (10, "기존 벽지 철거",        "-",               "㎡",  320,    1000, 1800,   200),
    (11, "욕실 방수",             "우레탄방수 2겹",   "㎡",   14,   15000,18000,  2000),
    (12, "하부장 600×720",        "W600×H720",       "EA",    6,   90000,75000,15000),
    (13, "상부장 600×350",        "W600×H350",       "EA",    5,   65000,45000,10000),
    # review_flag 5개
    (14, "인건비",                "-",               "식",    1,       0,480000,20000),
    (15, "EP도장 2차",            "에폭시",           "㎡",   20,   12000,14000,  2000),
    (16, "걸레받이 MDF 10T",      "MDF 10T",         "M",    95,    2500, 2500,   500),
    (17, "기타 잡공사",           "-",               "식",    1,       0,270000,30000),
    (18, "시스템 선반 W900",      "W900×D300",       "EA",    3,   40000,38000,  7000),
]

SECTION_ROWS = {
    "철거공사": [8, 10],
    "목공사": [1, 16],
    "도장/도배공사": [2, 7, 15],
    "바닥공사": [3, 11],
    "타일공사": [4, 5],
    "창호공사": [6],
    "주방공사": [12, 13],
    "폐기물처리": [9],
    "잡공사": [17, 18],
    "미분류(검토필요)": [14],
}


def set_cell(ws, row, col, value, font=None, fill=None, alignment=None,
             border=None, number_format=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if alignment:
        cell.alignment = alignment
    if border:
        cell.border = border
    if number_format:
        cell.number_format = number_format
    return cell


def create_demo_excel():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()

    # ── 시트 1: 세부내역서 ─────────────────────────────────────
    ws = wb.active
    ws.title = "세부내역서"

    # 열 너비 설정
    col_widths = {
        "A": 6,   # 번호
        "B": 30,  # 품명
        "C": 16,  # 규격
        "D": 6,   # 단위
        "E": 8,   # 수량
        "F": 12,  # 재료비단가
        "G": 12,  # 노무비단가
        "H": 10,  # 경비
        "I": 13,  # 단가
        "J": 14,  # 금액
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # 행 높이
    ws.row_dimensions[1].height = 8
    ws.row_dimensions[2].height = 8
    ws.row_dimensions[3].height = 36
    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 18
    ws.row_dimensions[6].height = 30

    # ── 헤더 영역 (행 1~6) ─────────────────────────────────────

    # 행 3: 타이틀
    ws.merge_cells("A3:J3")
    set_cell(ws, 3, 1, "한 샘 리 하 우 스  세 부 내 역 서",
             font=Font(name="맑은 고딕", bold=True, size=14),
             alignment=CENTER,
             fill=PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid"))
    ws["A3"].font = Font(name="맑은 고딕", bold=True, size=14, color="FFFFFF")

    # 행 4: 공사명
    ws.merge_cells("A4:E4")
    set_cell(ws, 4, 1, "공사명: ○○아파트 인테리어 공사 (34평)",
             font=Font(name="맑은 고딕", size=10),
             alignment=LEFT)
    ws.merge_cells("F4:J4")
    set_cell(ws, 4, 6, "현장주소: 서울 강남구 ○○동 ○○○호",
             font=Font(name="맑은 고딕", size=10),
             alignment=LEFT)

    # 행 5: 공사기간
    ws.merge_cells("A5:E5")
    set_cell(ws, 5, 1, "공사기간: 2026-05-01 ~ 2026-05-20",
             font=Font(name="맑은 고딕", size=10),
             alignment=LEFT)
    ws.merge_cells("F5:J5")
    set_cell(ws, 5, 6, "담당자: 홍길동 실장 / 010-1234-5678",
             font=Font(name="맑은 고딕", size=10),
             alignment=LEFT)

    # 행 6: 열 헤더
    headers = ["번호", "품명", "규격", "단위", "수량",
               "재료비\n단가", "노무비\n단가", "경비", "단가", "금액"]
    for col, h in enumerate(headers, 1):
        set_cell(ws, 6, col, h,
                 font=HEADER_FONT,
                 fill=HEADER_FILL,
                 alignment=CENTER,
                 border=HEADER_BORDER)

    # ── 데이터 행 (7행부터) ────────────────────────────────────
    data_row = 7
    for seq, name, spec, unit, qty, mat, labor, exp in ITEMS:
        # 홀짝 줄 배경 교차
        row_fill = WHITE_FILL if seq % 2 == 1 else LIGHT_GRAY

        set_cell(ws, data_row, 1, seq, BODY_FONT, row_fill, CENTER, THIN_BORDER)
        set_cell(ws, data_row, 2, name, BODY_FONT, row_fill, LEFT, THIN_BORDER)
        set_cell(ws, data_row, 3, spec, BODY_FONT, row_fill, CENTER, THIN_BORDER)
        set_cell(ws, data_row, 4, unit, BODY_FONT, row_fill, CENTER, THIN_BORDER)
        set_cell(ws, data_row, 5, qty, BODY_FONT, row_fill, RIGHT, THIN_BORDER, AREA_FORMAT)

        # 재료비/노무비/경비 단가 (수식 아님 — AI가 읽을 값)
        set_cell(ws, data_row, 6, mat, BODY_FONT, row_fill, RIGHT, THIN_BORDER, NUM_FORMAT)
        set_cell(ws, data_row, 7, labor, BODY_FONT, row_fill, RIGHT, THIN_BORDER, NUM_FORMAT)
        set_cell(ws, data_row, 8, exp, BODY_FONT, row_fill, RIGHT, THIN_BORDER, NUM_FORMAT)

        # 단가 = F+G+H (수식)
        unit_price_formula = f"=F{data_row}+G{data_row}+H{data_row}"
        set_cell(ws, data_row, 9, unit_price_formula,
                 BODY_FONT, row_fill, RIGHT, THIN_BORDER, NUM_FORMAT)

        # 금액 = E×I (수식)
        amount_formula = f"=E{data_row}*I{data_row}"
        set_cell(ws, data_row, 10, amount_formula,
                 BODY_FONT, row_fill, RIGHT, THIN_BORDER, NUM_FORMAT)

        ws.row_dimensions[data_row].height = 18
        data_row += 1

    # ── 합계 행 ────────────────────────────────────────────────
    total_row = data_row

    ws.merge_cells(f"A{total_row}:I{total_row}")
    set_cell(ws, total_row, 1, "합  계",
             font=TOTAL_FONT,
             fill=TOTAL_FILL,
             alignment=CENTER,
             border=HEADER_BORDER)

    total_formula = f"=SUM(J7:J{total_row - 1})"
    set_cell(ws, total_row, 10, total_formula,
             font=TOTAL_FONT,
             fill=TOTAL_FILL,
             alignment=RIGHT,
             border=HEADER_BORDER,
             number_format=NUM_FORMAT)

    ws.row_dimensions[total_row].height = 22

    # VAT 행
    vat_row = total_row + 1
    ws.merge_cells(f"A{vat_row}:I{vat_row}")
    set_cell(ws, vat_row, 1, "부가가치세 (10%)",
             font=BODY_FONT,
             fill=LIGHT_GRAY,
             alignment=CENTER,
             border=THIN_BORDER)
    set_cell(ws, vat_row, 10, f"=J{total_row}*0.1",
             font=BODY_FONT,
             fill=LIGHT_GRAY,
             alignment=RIGHT,
             border=THIN_BORDER,
             number_format=NUM_FORMAT)
    ws.row_dimensions[vat_row].height = 18

    # 총액 행
    grand_row = vat_row + 1
    ws.merge_cells(f"A{grand_row}:I{grand_row}")
    set_cell(ws, grand_row, 1, "총 공사금액",
             font=Font(name="맑은 고딕", bold=True, size=11),
             fill=PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid"),
             alignment=CENTER,
             border=HEADER_BORDER)
    ws[f"A{grand_row}"].font = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")

    set_cell(ws, grand_row, 10, f"=J{total_row}+J{vat_row}",
             font=Font(name="맑은 고딕", bold=True, size=11),
             fill=PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid"),
             alignment=RIGHT,
             border=HEADER_BORDER,
             number_format=NUM_FORMAT)
    ws[f"J{grand_row}"].font = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
    ws.row_dimensions[grand_row].height = 24

    # ── 시트 2: 표지 ───────────────────────────────────────────
    ws2 = wb.create_sheet("표지")
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 15
    ws2.column_dimensions["B"].width = 35

    ws2.merge_cells("A1:B1")
    ws2.row_dimensions[1].height = 50
    set_cell(ws2, 1, 1, "한 샘  리 하 우 스",
             font=Font(name="맑은 고딕", bold=True, size=20, color="FFFFFF"),
             fill=PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid"),
             alignment=CENTER)

    info = [
        ("공사명", "○○아파트 인테리어 공사"),
        ("현장주소", "서울 강남구 ○○동 ○○○호"),
        ("공사면적", "112.39㎡ (34평)"),
        ("공사기간", "2026-05-01 ~ 2026-05-20"),
        ("담당자", "홍길동 실장"),
        ("연락처", "010-1234-5678"),
    ]
    for i, (label, value) in enumerate(info, 3):
        ws2.row_dimensions[i].height = 24
        set_cell(ws2, i, 1, label,
                 font=Font(name="맑은 고딕", bold=True, size=10),
                 fill=SECTION_FILL,
                 alignment=CENTER,
                 border=THIN_BORDER)
        set_cell(ws2, i, 2, value,
                 font=Font(name="맑은 고딕", size=10),
                 alignment=LEFT,
                 border=THIN_BORDER)

    # ── 저장 ───────────────────────────────────────────────────
    wb.save(OUTPUT_FILE)
    print(f"[OK] 샘플 엑셀 생성 완료: {OUTPUT_FILE}")
    print(f"   - 시트: 세부내역서, 표지")
    print(f"   - 데이터 행: 7행 시작, 18개 품목")
    print(f"   - 합계 행: {total_row}행")
    print(f"   - 수식 셀: I열(단가=F+G+H), J열(금액=ExI), 합계/VAT/총액")
    print()
    print("   [자동분류 예상 13개]")
    for seq, name, *_ in ITEMS[:13]:
        print(f"   {seq:2d}. {name}")
    print()
    print("   [review_flag 예상 5개]")
    for seq, name, *_ in ITEMS[13:]:
        print(f"   {seq:2d}. {name}")


if __name__ == "__main__":
    create_demo_excel()
